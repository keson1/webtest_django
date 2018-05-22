#!/usr/bin/env python
# encoding: utf-8
'''
@author: Limz
@mail: limz@yisa.com
@name: batchimport.py
@time: 2018/5/20 23:09
@Description:
'''

from openpyxl import load_workbook
import os
def batchimport():
    basepath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(basepath, 'batch.xlsx')
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = ws.max_row
    nodelist = []
    if rows >= 2:
        for row in xrange(2, rows+1):
            node = {"ip": ws.cell(row=row, column=1).value,
                    "nodename": ws.cell(row=row, column=2).value,
                    "projectname":ws.cell(row=row, column=3).value,
                    "sshport": ws.cell(row=row, column=4).value,
                    "rootpwd": ws.cell(row=row, column=5).value,
                    "num": ws.cell(row=row, column=6).value,
                    "uptime": ws.cell(row=row, column=7).value,
                    "remark": ws.cell(row=row, column=8).value,
                    "position": ws.cell(row=row, column=9).value
            }
            nodelist.append(node)
    #print nodelist
    return nodelist